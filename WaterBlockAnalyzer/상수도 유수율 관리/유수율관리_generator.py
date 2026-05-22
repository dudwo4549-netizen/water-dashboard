# -*- coding: utf-8 -*-
"""
상수도 목표유수율 관리 엑셀 템플릿 생성기

출력:
  유수율관리_메인.xlsx       - openpyxl: 데이터 입력 / 유수율 현황 / 종합 현황
  유수율관리_차트리포트.xlsx  - xlsxwriter: 고품질 차트 전용 리포트
"""

import pathlib
import datetime
import openpyxl
import xlsxwriter
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 구성 상수
# ─────────────────────────────────────────────
ZONES = ["1구역", "2구역", "3구역", "4구역", "5구역", "6구역", "7구역", "8구역"]
NUM_ZONES = len(ZONES)
MONTHS = [f"{m}월" for m in range(1, 13)]
# 구역별 기본 목표유수율(%) — 실제 운영 시 O열에서 직접 수정
DEFAULT_TARGETS = [85.0, 86.0, 84.0, 87.0, 85.0, 83.0, 86.0, 85.0]
CF_BAND = 5          # 노랑 밴드: 목표 대비 -5%p ~ 0
CHART_YMIN = 50
CHART_YMAX = 100

OUT_DIR = pathlib.Path(__file__).parent
MAIN_FILE = OUT_DIR / "유수율관리_메인.xlsx"
CHART_FILE = OUT_DIR / "유수율관리_차트리포트.xlsx"

# 시트명
SH_INPUT = "데이터 입력"
SH_CALC  = "유수율 현황"
SH_DASH  = "종합 현황"

# ─────────────────────────────────────────────
# 스타일 헬퍼
# ─────────────────────────────────────────────
HEADER_BG    = "1F4E79"
SUBHDR_BG    = "2E75B6"
ODD_ROW_BG   = "DCE6F1"
TOTAL_BG     = "F2F2F2"
GREEN_CF     = "92D050"
YELLOW_CF    = "FFFF00"
RED_CF       = "FF0000"
ACHV_GREEN   = "C6EFCE"
FAIL_RED     = "FFCCCC"
FN           = "맑은 고딕"

def _font(bold=False, color="000000", size=10, name=FN):
    return Font(name=name, bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin", color="B8CCE4"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _right():
    return Alignment(horizontal="right", vertical="center")

def _apply_cell(cell, value=None, font=None, fill=None, border=None,
                alignment=None, number_format=None):
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format

def _style_header_row(ws, row, col_start, col_end, bg=HEADER_BG):
    fill = _fill(bg)
    font = _font(bold=True, color="FFFFFF", size=11)
    border = _border()
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = _center(wrap=True)

def _style_data_range(ws, min_row, max_row, min_col, max_col,
                      odd_fill=True, number_format=None):
    border = _border()
    for r in range(min_row, max_row + 1):
        row_fill = _fill(ODD_ROW_BG) if (odd_fill and r % 2 == 1) else _fill("FFFFFF")
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if cell.fill.patternType != "solid" or cell.fill.fgColor.rgb == "00000000":
                cell.fill = row_fill
            if number_format and cell.value is not None:
                cell.number_format = number_format

def _merge_title(ws, row, max_col, text, bg=HEADER_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    _apply_cell(
        cell, value=text,
        font=_font(bold=True, color="FFFFFF", size=14, name=FN),
        fill=_fill(bg),
        border=_border(),
        alignment=_center(),
    )
    ws.row_dimensions[row].height = 32


# ─────────────────────────────────────────────
# 크로스시트 수식 헬퍼 (공백 포함 시트명 → 작은따옴표)
# ─────────────────────────────────────────────
def _xref(sheet_name, cell_addr):
    if " " in sheet_name:
        return f"'{sheet_name}'!{cell_addr}"
    return f"{sheet_name}!{cell_addr}"


# ─────────────────────────────────────────────
# Sheet 1: "데이터 입력"
# ─────────────────────────────────────────────
def build_input_sheet(wb):
    ws = wb.create_sheet(SH_INPUT)
    ws.sheet_properties.tabColor = "1F4E79"

    MAX_COL = 15  # A~O

    # 행 레이아웃
    # Row 1: 타이틀
    # Row 2: 헤더 (급수구역 / 1월~12월 / 연간합계 / 목표유수율)
    # Row 3..18: 구역별 급수량(홀수행), 유수량(짝수행)  8구역×2=16행
    # Row 19: 구분선(빈행)
    # Row 20: 시스템 전체 급수량
    # Row 21: 시스템 전체 유수량

    # ── 타이틀
    _merge_title(ws, 1, MAX_COL, "상수도 유수율 관리 — 데이터 입력")

    # ── 헤더행
    headers = ["급수구역"] + MONTHS + ["연간합계", "목표유수율(%)"]
    for c, h in enumerate(headers, start=1):
        _apply_cell(
            ws.cell(row=2, column=c), value=h,
            font=_font(bold=True, color="FFFFFF", size=10),
            fill=_fill(SUBHDR_BG),
            border=_border(),
            alignment=_center(wrap=True),
        )
    ws.row_dimensions[2].height = 36

    # ── 구역별 데이터 행
    for i, zone in enumerate(ZONES):
        r_sup = 3 + i * 2      # 급수량 행
        r_rev = 3 + i * 2 + 1  # 유수량 행

        row_bg = _fill(ODD_ROW_BG) if i % 2 == 0 else _fill("FFFFFF")

        for r, label in [(r_sup, f"{zone} 급수량(㎥)"),
                         (r_rev, f"{zone} 유수량(㎥)")]:
            ws.row_dimensions[r].height = 20
            # A열: 라벨
            _apply_cell(
                ws.cell(row=r, column=1), value=label,
                font=_font(bold=(r == r_sup), size=10),
                fill=row_bg,
                border=_border(),
                alignment=_center(wrap=False),
            )
            # B~M: 입력 셀
            for col in range(2, 14):
                cell = ws.cell(row=r, column=col)
                cell.fill = row_bg
                cell.border = _border()
                cell.alignment = _right()
                cell.number_format = "#,##0"

            # N열: 연간합계 수식
            col_n = ws.cell(row=r, column=14)
            col_n.value = f"=SUM(B{r}:M{r})"
            col_n.font = _font(bold=True, size=10)
            col_n.fill = _fill(TOTAL_BG)
            col_n.border = _border()
            col_n.alignment = _right()
            col_n.number_format = "#,##0"

        # O열: 목표유수율 (급수량 행에만)
        o_cell = ws.cell(row=r_sup, column=15)
        o_cell.value = DEFAULT_TARGETS[i]
        o_cell.font = _font(bold=True, color="1F4E79", size=10)
        o_cell.fill = _fill("E2EFDA")
        o_cell.border = _border()
        o_cell.alignment = _center()
        o_cell.number_format = "0.0"
        # 유수량 행 O열: 빈칸 스타일만
        o_blank = ws.cell(row=r_rev, column=15)
        o_blank.fill = _fill("E2EFDA")
        o_blank.border = _border()

    # ── 시스템 전체 합계 행 (Row 20, 21)
    ws.row_dimensions[19].height = 8  # 구분선
    for r_total, label in [(20, "시스템 전체 급수량(㎥)"), (21, "시스템 전체 유수량(㎥)")]:
        ws.row_dimensions[r_total].height = 22
        _apply_cell(
            ws.cell(row=r_total, column=1), value=label,
            font=_font(bold=True, color="FFFFFF", size=10),
            fill=_fill(HEADER_BG),
            border=_border(),
            alignment=_center(),
        )
        for col in range(2, 15):
            col_letter = get_column_letter(col)
            # 급수량 행의 홀수 데이터행(r_sup)만 합산: 3,5,7,...,17
            if r_total == 20:
                sum_cells = "+".join(
                    f"{col_letter}{3 + k * 2}" for k in range(NUM_ZONES)
                )
            else:
                sum_cells = "+".join(
                    f"{col_letter}{3 + k * 2 + 1}" for k in range(NUM_ZONES)
                )
            if col == 14:  # N열: 전체 연간합계
                formula = f"=SUM(B{r_total}:M{r_total})"
            else:
                formula = f"={sum_cells}"
            cell = ws.cell(row=r_total, column=col)
            cell.value = formula
            cell.font = _font(bold=True, color="FFFFFF", size=10)
            cell.fill = _fill(HEADER_BG)
            cell.border = _border()
            cell.alignment = _right()
            cell.number_format = "#,##0"
        # O열 비워두기
        o_cell = ws.cell(row=r_total, column=15)
        o_cell.fill = _fill(HEADER_BG)
        o_cell.border = _border()

    # ── 열 너비
    ws.column_dimensions["A"].width = 24
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["N"].width = 12
    ws.column_dimensions["O"].width = 14

    # ── 틀 고정 & 인쇄
    ws.freeze_panes = "B3"
    ws.print_area = f"A1:O21"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    # ── 데이터 검증: 입력 셀 (급수량/유수량) — 0 이상
    dv_data = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="입력 오류",
        error="0 이상의 숫자를 입력하세요.",
        promptTitle="급수량/유수량 입력",
        prompt="단위: ㎥, 0 이상 숫자",
    )
    dv_data.sqref = "B3:M18"
    ws.add_data_validation(dv_data)

    # ── 데이터 검증: 목표유수율 0~100
    dv_target = DataValidation(
        type="decimal", operator="between", formula1="0", formula2="100",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="목표유수율 오류",
        error="0~100 사이 값을 입력하세요.",
        promptTitle="목표유수율 입력",
        prompt="단위: %, 예) 85.0",
    )
    for k in range(NUM_ZONES):
        dv_target.add(f"O{3 + k * 2}")
    ws.add_data_validation(dv_target)

    return ws


# ─────────────────────────────────────────────
# Sheet 2: "유수율 현황"
# ─────────────────────────────────────────────
def build_calc_sheet(wb):
    ws = wb.create_sheet(SH_CALC)
    ws.sheet_properties.tabColor = "2E75B6"

    MAX_COL = 15

    _merge_title(ws, 1, MAX_COL, "상수도 유수율 현황")

    headers = ["급수구역"] + MONTHS + ["연간평균(%)", "목표유수율(%)"]
    for c, h in enumerate(headers, start=1):
        _apply_cell(
            ws.cell(row=2, column=c), value=h,
            font=_font(bold=True, color="FFFFFF", size=10),
            fill=_fill(SUBHDR_BG),
            border=_border(),
            alignment=_center(wrap=True),
        )
    ws.row_dimensions[2].height = 36

    # ── 구역별 유수율 수식
    for i, zone in enumerate(ZONES):
        calc_row = 3 + i           # 유수율현황 시트의 해당 행
        r_sup = 3 + i * 2          # 데이터입력 시트의 급수량 행
        r_rev = r_sup + 1

        row_bg = _fill(ODD_ROW_BG) if i % 2 == 0 else _fill("FFFFFF")

        # A열: 구역명
        _apply_cell(
            ws.cell(row=calc_row, column=1), value=zone,
            font=_font(bold=True, size=10),
            fill=row_bg,
            border=_border(),
            alignment=_center(),
        )
        ws.row_dimensions[calc_row].height = 22

        # B~M: 유수율 수식
        for m_idx, col in enumerate(range(2, 14)):
            col_letter = get_column_letter(col)
            sup_ref = _xref(SH_INPUT, f"{col_letter}{r_sup}")
            rev_ref = _xref(SH_INPUT, f"{col_letter}{r_rev}")
            formula = f"=IF({sup_ref}=0,\"\",{rev_ref}/{sup_ref}*100)"
            cell = ws.cell(row=calc_row, column=col)
            cell.value = formula
            cell.fill = row_bg
            cell.border = _border()
            cell.alignment = _center()
            cell.number_format = "0.0"

        # N열: 연간평균 (데이터 없는 월 제외)
        n_cell = ws.cell(row=calc_row, column=14)
        n_cell.value = f'=IFERROR(AVERAGEIF(B{calc_row}:M{calc_row},"<>"""),"")'
        n_cell.font = _font(bold=True, size=10)
        n_cell.fill = _fill(TOTAL_BG)
        n_cell.border = _border()
        n_cell.alignment = _center()
        n_cell.number_format = "0.0"

        # O열: 목표유수율 참조
        o_cell = ws.cell(row=calc_row, column=15)
        o_cell.value = f"={_xref(SH_INPUT, f'O{r_sup}')}"
        o_cell.font = _font(bold=True, color="1F4E79", size=10)
        o_cell.fill = _fill("E2EFDA")
        o_cell.border = _border()
        o_cell.alignment = _center()
        o_cell.number_format = "0.0"

    # ── 구분 행
    ws.row_dimensions[NUM_ZONES + 3].height = 8

    # ── 시스템 전체 유수율 행 (row 12)
    total_row = NUM_ZONES + 4   # = 12
    ws.row_dimensions[total_row].height = 22
    _apply_cell(
        ws.cell(row=total_row, column=1), value="시스템 전체",
        font=_font(bold=True, color="FFFFFF", size=10),
        fill=_fill(HEADER_BG),
        border=_border(),
        alignment=_center(),
    )
    for col in range(2, 14):
        col_letter = get_column_letter(col)
        sup_parts = "+".join(
            _xref(SH_INPUT, f"{col_letter}{3 + k * 2}") for k in range(NUM_ZONES)
        )
        rev_parts = "+".join(
            _xref(SH_INPUT, f"{col_letter}{3 + k * 2 + 1}") for k in range(NUM_ZONES)
        )
        formula = f"=IF(({sup_parts})=0,\"\",({rev_parts})/({sup_parts})*100)"
        cell = ws.cell(row=total_row, column=col)
        cell.value = formula
        cell.font = _font(bold=True, color="FFFFFF", size=10)
        cell.fill = _fill(HEADER_BG)
        cell.border = _border()
        cell.alignment = _center()
        cell.number_format = "0.0"

    # N열: 시스템 전체 연간평균
    n_total = ws.cell(row=total_row, column=14)
    n_total.value = f'=IFERROR(AVERAGEIF(B{total_row}:M{total_row},"<>"""),"")'
    n_total.font = _font(bold=True, color="FFFFFF", size=10)
    n_total.fill = _fill(HEADER_BG)
    n_total.border = _border()
    n_total.alignment = _center()
    n_total.number_format = "0.0"

    # O열 비워두기
    o_total = ws.cell(row=total_row, column=15)
    o_total.fill = _fill(HEADER_BG)
    o_total.border = _border()

    # ── 조건부서식 (B3:M10 — 각 구역×월)
    cf_range = f"B3:M{2 + NUM_ZONES}"

    green_fill  = _fill(GREEN_CF)
    yellow_fill = _fill(YELLOW_CF)
    red_fill    = _fill(RED_CF)

    # 수식은 범위의 첫 셀(B3) 기준으로 작성; $O3: 절대열·상대행
    ws.conditional_formatting.add(
        cf_range,
        FormulaRule(
            formula=['AND(B3<>"",B3>=$O3)'],
            fill=green_fill,
            stopIfTrue=True,
        )
    )
    ws.conditional_formatting.add(
        cf_range,
        FormulaRule(
            formula=[f'AND(B3<>"",B3>=$O3-{CF_BAND},B3<$O3)'],
            fill=yellow_fill,
            stopIfTrue=True,
        )
    )
    ws.conditional_formatting.add(
        cf_range,
        FormulaRule(
            formula=[f'AND(B3<>"",B3<$O3-{CF_BAND})'],
            fill=red_fill,
            stopIfTrue=True,
        )
    )

    # N열 연간평균에도 조건부서식
    n_cf_range = f"N3:N{2 + NUM_ZONES}"
    ws.conditional_formatting.add(
        n_cf_range,
        FormulaRule(formula=['AND(N3<>"",N3>=$O3)'],       fill=green_fill,  stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        n_cf_range,
        FormulaRule(formula=[f'AND(N3<>"",N3>=$O3-{CF_BAND},N3<$O3)'], fill=yellow_fill, stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        n_cf_range,
        FormulaRule(formula=[f'AND(N3<>"",N3<$O3-{CF_BAND})'],         fill=red_fill,    stopIfTrue=True)
    )

    # ── 열 너비
    ws.column_dimensions["A"].width = 14
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions["N"].width = 13
    ws.column_dimensions["O"].width = 14

    # ── 틀 고정 & 인쇄
    ws.freeze_panes = "B3"
    ws.print_area = f"A1:O{total_row}"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    return ws, total_row


# ─────────────────────────────────────────────
# Sheet 3: "종합 현황"
# ─────────────────────────────────────────────
def build_dashboard_sheet(wb, calc_total_row):
    ws = wb.create_sheet(SH_DASH)
    ws.sheet_properties.tabColor = "70AD47"

    MAX_COL = 14

    _merge_title(ws, 1, MAX_COL, "상수도 유수율 종합 현황")

    # ── [ 섹션 1 ] 월별 유수율 추이 테이블 ──────────────────────
    # Row 3: 섹션 타이틀
    ws.merge_cells("A3:F3")
    _apply_cell(
        ws.cell(row=3, column=1), value="▶ 월별 유수율 추이 (시스템 전체)",
        font=_font(bold=True, color="FFFFFF", size=11),
        fill=_fill(SUBHDR_BG),
        border=_border(),
        alignment=_center(),
    )
    ws.row_dimensions[3].height = 24

    # Row 4: 헤더
    for c, h in enumerate(["월", "유수율(%)", "목표유수율(%)"], start=1):
        _apply_cell(
            ws.cell(row=4, column=c), value=h,
            font=_font(bold=True, color="FFFFFF", size=10),
            fill=_fill(HEADER_BG),
            border=_border(),
            alignment=_center(wrap=True),
        )
    ws.row_dimensions[4].height = 30

    # Row 5~16: 1월~12월 데이터
    for m_idx, month in enumerate(MONTHS):
        r = 5 + m_idx
        col_letter = get_column_letter(m_idx + 2)  # B~M
        calc_month_rate = _xref(SH_CALC, f"{col_letter}{calc_total_row}")
        # 목표: 전체 구역 목표 평균
        target_avg = f'=IFERROR(AVERAGE({_xref(SH_CALC, f"O3:O{2+NUM_ZONES}")}),"")'

        ws.row_dimensions[r].height = 20
        row_bg = _fill(ODD_ROW_BG) if m_idx % 2 == 0 else _fill("FFFFFF")

        _apply_cell(ws.cell(row=r, column=1), value=month,
                    font=_font(bold=True, size=10), fill=row_bg,
                    border=_border(), alignment=_center())
        b_cell = ws.cell(row=r, column=2)
        b_cell.value = f"={calc_month_rate}"
        b_cell.fill = row_bg; b_cell.border = _border()
        b_cell.alignment = _center(); b_cell.number_format = "0.0"
        c_cell = ws.cell(row=r, column=3)
        c_cell.value = target_avg
        c_cell.fill = row_bg; c_cell.border = _border()
        c_cell.alignment = _center(); c_cell.number_format = "0.0"

    # 월별 조건부서식 (B5:B16)
    ws.conditional_formatting.add(
        "B5:B16",
        FormulaRule(formula=['AND(B5<>"",B5>=$C5)'],                       fill=_fill(GREEN_CF),  stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        "B5:B16",
        FormulaRule(formula=[f'AND(B5<>"",B5>=$C5-{CF_BAND},B5<$C5)'],    fill=_fill(YELLOW_CF), stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        "B5:B16",
        FormulaRule(formula=[f'AND(B5<>"",B5<$C5-{CF_BAND})'],             fill=_fill(RED_CF),    stopIfTrue=True)
    )

    # ── [ 섹션 2 ] 구역별 달성 현황 테이블 ──────────────────────
    # Row 18: 섹션 타이틀
    ws.merge_cells("A18:G18")
    _apply_cell(
        ws.cell(row=18, column=1), value="▶ 급수구역별 연간 달성 현황",
        font=_font(bold=True, color="FFFFFF", size=11),
        fill=_fill(SUBHDR_BG),
        border=_border(),
        alignment=_center(),
    )
    ws.row_dimensions[18].height = 24

    # Row 19: 헤더
    zone_headers = ["급수구역", "연간평균(%)", "목표유수율(%)", "달성여부", "목표대비(%p)"]
    for c, h in enumerate(zone_headers, start=1):
        _apply_cell(
            ws.cell(row=19, column=c), value=h,
            font=_font(bold=True, color="FFFFFF", size=10),
            fill=_fill(HEADER_BG),
            border=_border(),
            alignment=_center(wrap=True),
        )
    ws.row_dimensions[19].height = 30

    # Row 20~27: 구역별
    for i, zone in enumerate(ZONES):
        r = 20 + i
        calc_row = 3 + i
        avg_ref   = _xref(SH_CALC, f"N{calc_row}")
        target_ref = _xref(SH_CALC, f"O{calc_row}")
        row_bg = _fill(ODD_ROW_BG) if i % 2 == 0 else _fill("FFFFFF")
        ws.row_dimensions[r].height = 20

        _apply_cell(ws.cell(row=r, column=1), value=zone,
                    font=_font(bold=True, size=10), fill=row_bg,
                    border=_border(), alignment=_center())
        b_cell = ws.cell(row=r, column=2)
        b_cell.value = f"={avg_ref}"
        b_cell.fill = row_bg; b_cell.border = _border()
        b_cell.alignment = _center(); b_cell.number_format = "0.0"
        c_cell = ws.cell(row=r, column=3)
        c_cell.value = f"={target_ref}"
        c_cell.fill = row_bg; c_cell.border = _border()
        c_cell.alignment = _center(); c_cell.number_format = "0.0"
        d_cell = ws.cell(row=r, column=4)
        d_cell.value = f'=IF(B{r}="","",IF(B{r}>=C{r},"달성","미달"))'
        d_cell.fill = row_bg; d_cell.border = _border(); d_cell.alignment = _center()
        d_cell.font = _font(bold=True, size=10)
        e_cell = ws.cell(row=r, column=5)
        e_cell.value = f'=IF(B{r}="","",B{r}-C{r})'
        e_cell.fill = row_bg; e_cell.border = _border()
        e_cell.alignment = _center(); e_cell.number_format = "+0.0;-0.0;0.0"

    # Row 28: 시스템 전체
    r_sys = 28
    ws.row_dimensions[r_sys].height = 22
    sys_avg_ref    = _xref(SH_CALC, f"N{calc_total_row}")
    sys_target_ref = f'=IFERROR(AVERAGE({_xref(SH_CALC, f"O3:O{2+NUM_ZONES}")}),"")'
    _apply_cell(ws.cell(row=r_sys, column=1), value="시스템 전체",
                font=_font(bold=True, color="FFFFFF"), fill=_fill(HEADER_BG),
                border=_border(), alignment=_center())
    b28 = ws.cell(row=r_sys, column=2)
    b28.value = f"={sys_avg_ref}"; b28.font = _font(bold=True, color="FFFFFF")
    b28.fill = _fill(HEADER_BG); b28.border = _border()
    b28.alignment = _center(); b28.number_format = "0.0"
    c28 = ws.cell(row=r_sys, column=3)
    c28.value = sys_target_ref; c28.font = _font(bold=True, color="FFFFFF")
    c28.fill = _fill(HEADER_BG); c28.border = _border()
    c28.alignment = _center(); c28.number_format = "0.0"
    d28 = ws.cell(row=r_sys, column=4)
    d28.value = f'=IF(B{r_sys}="","",IF(B{r_sys}>=C{r_sys},"달성","미달"))'
    d28.font = _font(bold=True, color="FFFFFF"); d28.fill = _fill(HEADER_BG)
    d28.border = _border(); d28.alignment = _center()
    e28 = ws.cell(row=r_sys, column=5)
    e28.value = f'=IF(B{r_sys}="","",B{r_sys}-C{r_sys})'
    e28.font = _font(bold=True, color="FFFFFF"); e28.fill = _fill(HEADER_BG)
    e28.border = _border(); e28.alignment = _center(); e28.number_format = "+0.0;-0.0;0.0"

    # 달성여부 조건부서식 (D20:D27)
    ws.conditional_formatting.add(
        "D20:D27",
        FormulaRule(formula=['D20="달성"'], fill=_fill(ACHV_GREEN), stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        "D20:D27",
        FormulaRule(formula=['D20="미달"'], fill=_fill(FAIL_RED), stopIfTrue=True)
    )
    # 연간평균 조건부서식 (B20:B27)
    ws.conditional_formatting.add(
        "B20:B27",
        FormulaRule(formula=['AND(B20<>"",B20>=$C20)'],                        fill=_fill(GREEN_CF),  stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        "B20:B27",
        FormulaRule(formula=[f'AND(B20<>"",B20>=$C20-{CF_BAND},B20<$C20)'],   fill=_fill(YELLOW_CF), stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        "B20:B27",
        FormulaRule(formula=[f'AND(B20<>"",B20<$C20-{CF_BAND})'],              fill=_fill(RED_CF),    stopIfTrue=True)
    )

    # ── [ 차트 1 ] 월별 유수율 추이 꺾은선 차트 (openpyxl)
    chart_line = LineChart()
    chart_line.title = "월별 유수율 추이 (시스템 전체)"
    chart_line.style = 10
    chart_line.y_axis.title = "유수율 (%)"
    chart_line.x_axis.title = "월"
    chart_line.y_axis.numFmt = "0.0"
    chart_line.y_axis.scaling.min = CHART_YMIN
    chart_line.y_axis.scaling.max = CHART_YMAX
    chart_line.height = 14
    chart_line.width = 20

    data_ref = Reference(ws, min_col=2, max_col=3, min_row=4, max_row=16)
    chart_line.add_data(data_ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=5, max_row=16)
    chart_line.set_categories(cats)

    # 목표선 스타일: 빨간 파선
    chart_line.series[0].graphicalProperties.line.solidFill = "2E75B6"
    chart_line.series[0].graphicalProperties.line.width = 20000
    chart_line.series[1].graphicalProperties.line.solidFill = "FF0000"
    chart_line.series[1].graphicalProperties.line.width = 15000
    chart_line.series[1].graphicalProperties.line.dashDot = "dash"

    ws.add_chart(chart_line, "A31")

    # ── [ 차트 2 ] 구역별 연간 유수율 막대 차트 (openpyxl)
    chart_bar = BarChart()
    chart_bar.type = "col"
    chart_bar.title = "급수구역별 연간 유수율 달성 현황"
    chart_bar.style = 10
    chart_bar.y_axis.title = "유수율 (%)"
    chart_bar.y_axis.numFmt = "0.0"
    chart_bar.y_axis.scaling.min = CHART_YMIN
    chart_bar.y_axis.scaling.max = CHART_YMAX
    chart_bar.height = 14
    chart_bar.width = 20

    bar_data = Reference(ws, min_col=2, max_col=3, min_row=19, max_row=27)
    chart_bar.add_data(bar_data, titles_from_data=True)
    zone_cats = Reference(ws, min_col=1, min_row=20, max_row=27)
    chart_bar.set_categories(zone_cats)

    chart_bar.series[0].graphicalProperties.solidFill = "2E75B6"
    chart_bar.series[1].graphicalProperties.line.solidFill = "FF0000"

    ws.add_chart(chart_bar, "J31")

    # ── 열 너비
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    # ── 범례 설명 (Row 29~30)
    ws.row_dimensions[29].height = 8
    ws.merge_cells("A30:E30")
    legend_cell = ws.cell(row=30, column=1)
    legend_cell.value = (
        f"색상 기준: ■ 초록(달성, 유수율≥목표)  "
        f"■ 노랑(근접, 목표-{CF_BAND}%p≤유수율<목표)  "
        f"■ 빨강(미달, 유수율<목표-{CF_BAND}%p)"
    )
    legend_cell.font = _font(size=9, color="595959")
    legend_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[30].height = 18

    ws.freeze_panes = "B5"

    return ws


# ─────────────────────────────────────────────
# xlsxwriter 차트 리포트 파일
# ─────────────────────────────────────────────
def build_chart_report(sample_data: dict):
    """
    sample_data = {
        "monthly_rates":  [float or None] × 12,  # 시스템 전체 월별 유수율(%)
        "monthly_target": float,                  # 목표유수율 (전체 평균)
        "zone_rates":     [float or None] × 8,    # 구역별 연간평균
        "zone_targets":   [float] × 8,            # 구역별 목표
        "zones":          [str] × 8,
        "months":         [str] × 12,
    }
    """
    wb = xlsxwriter.Workbook(str(CHART_FILE))
    ws = wb.add_worksheet("차트 리포트")

    # ── 포맷 정의
    fmt_title  = wb.add_format({"bold": True, "font_size": 16, "font_name": "맑은 고딕",
                                 "font_color": "FFFFFF", "bg_color": "1F4E79",
                                 "align": "center", "valign": "vcenter"})
    fmt_header = wb.add_format({"bold": True, "font_size": 10, "font_name": "맑은 고딕",
                                 "font_color": "FFFFFF", "bg_color": "2E75B6",
                                 "align": "center", "valign": "vcenter",
                                 "border": 1})
    fmt_data   = wb.add_format({"font_size": 10, "font_name": "맑은 고딕",
                                 "num_format": "0.0", "align": "center",
                                 "border": 1, "bg_color": "DCE6F1"})
    fmt_data2  = wb.add_format({"font_size": 10, "font_name": "맑은 고딕",
                                 "num_format": "0.0", "align": "center",
                                 "border": 1, "bg_color": "FFFFFF"})
    fmt_target = wb.add_format({"font_size": 10, "font_name": "맑은 고딕",
                                 "num_format": "0.0", "align": "center",
                                 "border": 1, "bg_color": "E2EFDA",
                                 "bold": True, "font_color": "375623"})
    fmt_section = wb.add_format({"bold": True, "font_size": 11, "font_name": "맑은 고딕",
                                  "font_color": "FFFFFF", "bg_color": "2E75B6",
                                  "align": "left", "valign": "vcenter"})

    ws.set_row(0, 36)
    ws.merge_range("A1:D1", "상수도 유수율 관리 — 차트 리포트", fmt_title)
    ws.set_column("A:A", 14)
    ws.set_column("B:B", 14)
    ws.set_column("C:C", 14)

    # ── 월별 데이터 테이블 (행 3~16)
    ws.set_row(2, 24)
    ws.merge_range("A3:C3", "▶ 월별 유수율 추이 (시스템 전체)", fmt_section)
    ws.set_row(3, 28)
    ws.write_row("A4", ["월", "유수율(%)", "목표유수율(%)"], fmt_header)

    months  = sample_data["months"]
    m_rates = sample_data["monthly_rates"]
    m_tgt   = sample_data["monthly_target"]

    for m_idx, month in enumerate(months):
        r = 4 + m_idx
        ws.set_row(r, 20)
        fmt_d = fmt_data if m_idx % 2 == 0 else fmt_data2
        ws.write(r, 0, month, fmt_d)
        val = m_rates[m_idx]
        ws.write(r, 1, val if val is not None else "", fmt_d)
        ws.write(r, 2, m_tgt, fmt_target)

    # ── 구역별 데이터 테이블 (행 18~27)
    ws.set_row(17, 24)
    ws.merge_range("A18:D18", "▶ 급수구역별 연간 달성 현황", fmt_section)
    ws.set_row(18, 28)
    ws.set_column("D:D", 14)
    ws.write_row("A19", ["급수구역", "연간평균(%)", "목표유수율(%)", "목표대비(%p)"], fmt_header)

    zones       = sample_data["zones"]
    z_rates     = sample_data["zone_rates"]
    z_targets   = sample_data["zone_targets"]

    for z_idx, zone in enumerate(zones):
        r = 19 + z_idx
        ws.set_row(r, 20)
        fmt_d = fmt_data if z_idx % 2 == 0 else fmt_data2
        ws.write(r, 0, zone, fmt_d)
        val = z_rates[z_idx]
        ws.write(r, 1, val if val is not None else "", fmt_d)
        ws.write(r, 2, z_targets[z_idx], fmt_target)
        diff = (val - z_targets[z_idx]) if val is not None else ""
        ws.write(r, 3, diff, fmt_d)

    # ── 차트 1: 월별 유수율 추이 꺾은선 + 목표선
    chart_line = wb.add_chart({"type": "line"})
    chart_line.set_title({"name": "월별 유수율 추이 (시스템 전체)",
                           "name_font": {"name": "맑은 고딕", "size": 13, "bold": True}})
    chart_line.set_x_axis({"name": "월", "name_font": {"name": "맑은 고딕"}})
    chart_line.set_y_axis({
        "name": "유수율 (%)",
        "name_font": {"name": "맑은 고딕"},
        "min": CHART_YMIN,
        "max": CHART_YMAX,
        "num_format": "0.0",
    })
    chart_line.set_size({"width": 480, "height": 320})
    chart_line.set_style(10)
    chart_line.set_legend({"position": "bottom", "font": {"name": "맑은 고딕"}})

    # 유수율 시리즈 (B열)
    chart_line.add_series({
        "name":       ["차트 리포트", 3, 1],
        "categories": ["차트 리포트", 4, 0, 15, 0],
        "values":     ["차트 리포트", 4, 1, 15, 1],
        "line":       {"color": "#2E75B6", "width": 2.5},
        "marker":     {"type": "circle", "size": 5,
                       "fill": {"color": "#2E75B6"},
                       "border": {"color": "#2E75B6"}},
    })
    # 목표 시리즈 (C열) — 빨간 파선
    chart_line.add_series({
        "name":       ["차트 리포트", 3, 2],
        "categories": ["차트 리포트", 4, 0, 15, 0],
        "values":     ["차트 리포트", 4, 2, 15, 2],
        "line":       {"color": "#FF0000", "width": 1.75, "dash_type": "dash"},
    })
    ws.insert_chart("F3", chart_line)

    # ── 차트 2: 구역별 막대 + 목표 꺾은선 혼합 차트
    chart_bar = wb.add_chart({"type": "column"})
    chart_bar.set_title({"name": "급수구역별 연간 유수율 달성 현황",
                          "name_font": {"name": "맑은 고딕", "size": 13, "bold": True}})
    chart_bar.set_x_axis({"name": "급수구역", "name_font": {"name": "맑은 고딕"}})
    chart_bar.set_y_axis({
        "name": "유수율 (%)",
        "name_font": {"name": "맑은 고딕"},
        "min": CHART_YMIN,
        "max": CHART_YMAX,
        "num_format": "0.0",
    })
    chart_bar.set_size({"width": 480, "height": 320})
    chart_bar.set_style(10)
    chart_bar.set_legend({"position": "bottom", "font": {"name": "맑은 고딕"}})

    chart_bar.add_series({
        "name":       ["차트 리포트", 18, 1],
        "categories": ["차트 리포트", 19, 0, 26, 0],
        "values":     ["차트 리포트", 19, 1, 26, 1],
        "fill":       {"color": "#2E75B6"},
        "gap":        100,
    })

    # 목표 꺾은선을 혼합 차트로 추가
    chart_line2 = wb.add_chart({"type": "line"})
    chart_line2.add_series({
        "name":       ["차트 리포트", 18, 2],
        "categories": ["차트 리포트", 19, 0, 26, 0],
        "values":     ["차트 리포트", 19, 2, 26, 2],
        "line":       {"color": "#FF0000", "width": 2, "dash_type": "dash"},
        "marker":     {"type": "diamond", "size": 6,
                       "fill": {"color": "#FF0000"},
                       "border": {"color": "#FF0000"}},
    })
    chart_bar.combine(chart_line2)
    ws.insert_chart("F18", chart_bar)

    wb.close()


# ─────────────────────────────────────────────
# main
# ─────────────────────────────────────────────
def main():
    # ── openpyxl: 메인 파일
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 기본 Sheet 제거

    build_input_sheet(wb)
    _, calc_total_row = build_calc_sheet(wb)
    build_dashboard_sheet(wb, calc_total_row)

    wb.save(str(MAIN_FILE))
    print(f"[완료] 메인 파일: {MAIN_FILE}")

    # ── xlsxwriter: 차트 리포트 파일 (샘플 데이터로 미리채움)
    # 실제 운영 시 DB/CSV에서 읽어 이 딕셔너리에 채워넣으면 됨
    sample_monthly = [82.1, 83.5, 84.0, 85.2, 86.1, 87.3,
                      88.0, 87.5, 86.8, 85.9, 84.7, 83.2]
    sample_zones   = [85.3, 86.1, 84.7, 87.2, 84.9, 83.5, 86.4, 85.8]

    build_chart_report({
        "monthly_rates":  sample_monthly,
        "monthly_target": sum(DEFAULT_TARGETS) / len(DEFAULT_TARGETS),
        "zone_rates":     sample_zones,
        "zone_targets":   DEFAULT_TARGETS,
        "zones":          ZONES,
        "months":         MONTHS,
    })
    print(f"[완료] 차트 리포트: {CHART_FILE}")
    print()
    print("사용 방법:")
    print("  1. '유수율관리_메인.xlsx' → '데이터 입력' 시트에 급수량/유수량 입력")
    print("  2. '유수율 현황' 시트에서 색상 자동 반영 확인")
    print("  3. '종합 현황' 시트에서 추이 차트 및 달성 현황 확인")
    print("  4. '유수율관리_차트리포트.xlsx' - xlsxwriter 고품질 차트 리포트")


if __name__ == "__main__":
    main()
